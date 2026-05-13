import openpyxl, json, re, sys
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r'D:\biju project.xlsx'
OUTPUT_JSON = r'c:\Projects\Billing\cable-bill-buddy-main\public\import_data.json'

BILL_SHEETS   = ['BILL-7','BILL-8','BILL-9','BILL-10','BILL-11','BILL-12']
MONTH_MAP     = {'BILL-7': ('2023','07'), 'BILL-8': ('2023','08'), 'BILL-9': ('2023','09'),
                 'BILL-10': ('2023','10'), 'BILL-11': ('2023','11'), 'BILL-12': ('2023','12')}

# ── helpers ──────────────────────────────────────────────────────────────────
def clean(v):
    if v is None: return ''
    s = str(v).replace('_x000D_', '').replace('\r', '').replace('\n', '').replace('\ufeff', '').strip()
    return s

def to_num(v):
    s = clean(v)
    try: return float(s)
    except: return 0.0

def is_data_row(row):
    """Row 0 is S.No field — check if first cell is a number"""
    try:
        int(float(str(row[0])))
        return True
    except:
        return False

def infer_plan(amount):
    """Map billing amount to plan id"""
    amt = to_num(amount)
    if amt <= 0:   return 'p1'
    if amt <= 330: return 'p1'  # Basic 200-330
    if amt <= 380: return 'p2'  # Standard HD 380
    if amt <= 450: return 'p4'  # Family Combo 405-450
    return 'p3'                  # Premium Sports 550+

def make_id(prefix, seed):
    import hashlib
    h = hashlib.md5(seed.encode()).hexdigest()[:9]
    return f"{prefix}{h}"

# ── load workbook ────────────────────────────────────────────────────────────
print("Loading workbook (may take a moment)...")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
print(f"Sheets: {wb.sheetnames}")

# ── collect all subscriber records ──────────────────────────────────────────
# Key = (name_clean, area_clean) → tracks latest record per customer
subscribers_map = {}   # key -> subscriber dict
payments_list   = []   # all payment records
invoices_list   = []   # all invoice records

for sheet_name in BILL_SHEETS:
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    yr, mo = MONTH_MAP[sheet_name]
    bill_date = f"{yr}-{mo}-01"
    due_date  = f"{yr}-{mo}-28"

    print(f"\nProcessing {sheet_name}...")
    count = 0

    for row in ws.iter_rows(values_only=True):
        if not is_data_row(row):
            continue

        sno          = clean(row[0])
        name         = clean(row[1])
        area         = clean(row[2])
        billing_amt  = clean(row[4])   # BILLING amount
        open_debit   = to_num(row[5])  # OPENING DEBIT
        open_credit  = to_num(row[6])  # OPENING CREDIT
        received     = to_num(row[7])  # RECEIVED
        close_debit  = to_num(row[10]) # CLOSING DEBIT
        close_credit = to_num(row[11]) # CLOSING CREDIT
        paid_status  = clean(row[12])  # True/False
        remark       = clean(row[13])  # Paid/Pending

        if not name:
            continue

        key = f"{name}|{area}"
        billing_num = to_num(billing_amt) if billing_amt not in ('OF', '', '0') else 0
        plan_id     = infer_plan(billing_num)

        # Balance: negative means customer owes, positive means advance
        balance = close_credit - close_debit

        # Subscriber record (update with latest month's data)
        if key not in subscribers_map:
            sub_id = make_id('SUB', key)
            subscribers_map[key] = {
                'id':          sub_id,
                'code':        f"STR{str(sno).zfill(4)}",
                'name':        name,
                'phone':       '',           
                'area':        area,
                'stbNumber':   '',           
                'planId':      'p1',         # Default plan
                'status':      'active',
                'expiryDate':  datetime.now().strftime("%Y-%m-%d"),
                'balance':     0,
                'autoBilling': True,
                'unpaidMonths': []
            }
        count += 1

    print(f"  → Processed {count} data rows")

# ── finalize subscribers ─────────────────────────────────────────────────────
subscribers_list = list(subscribers_map.values())
payments_list = [] # User requested only name and area
invoices_list = [] # User requested only name and area

print(f"\nTotal unique subscribers: {len(subscribers_list)}")
print(f"Total payments:           {len(payments_list)}")
print(f"Total invoices:           {len(invoices_list)}")

# Count missing fields
missing_phone  = sum(1 for s in subscribers_list if not s['phone'])
missing_stb    = sum(1 for s in subscribers_list if not s['stbNumber'])
print(f"Missing phone numbers:    {missing_phone}")
print(f"Missing STB numbers:      {missing_stb}")

# ── save output ──────────────────────────────────────────────────────────────
output = {
    'subscribers': subscribers_list,
    'payments':    payments_list,
    'invoices':    invoices_list,
    'expenses':    [],
    'reminders':   [],
    'agents':      [],
    'importedAt':  datetime.now().isoformat(),
    'source':      'D:\\biju project.xlsx',
    'stats': {
        'totalSubscribers': len(subscribers_list),
        'totalPayments':    len(payments_list),
        'totalInvoices':    len(invoices_list),
        'missingPhone':     missing_phone,
        'missingSTB':       missing_stb,
    }
}

with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n✅ Saved to: {OUTPUT_JSON}")

# Print sample subscribers
print("\n── Sample Subscribers ──")
for s in subscribers_list[:10]:
    print(f"  {s['code']} | {s['name']} | {s['area']} | Plan: {s['planId']} | Balance: {s['balance']} | Status: {s['status']}")
